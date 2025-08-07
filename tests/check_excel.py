#!/usr/bin/env python3
"""
Script to check the Excel file structure
"""

import openpyxl

def check_excel_file():
    """Check the structure of the generated Excel file"""
    try:
        # Load the Excel file
        wb = openpyxl.load_workbook('sbom_sample.xlsx')
        
        print(f"Excel file loaded successfully!")
        print(f"Number of worksheets: {len(wb.sheetnames)}")
        print(f"Worksheet names: {wb.sheetnames}")
        
        # Check each worksheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\nWorksheet: '{sheet_name}'")
            print(f"  - Number of rows: {ws.max_row}")
            print(f"  - Number of columns: {ws.max_column}")
            
            # Show first few rows
            print(f"  - First 3 rows:")
            for i in range(1, min(4, ws.max_row + 1)):
                row_data = []
                for j in range(1, min(10, ws.max_column + 1)):  # Show first 10 columns
                    cell_value = ws.cell(row=i, column=j).value
                    row_data.append(str(cell_value) if cell_value is not None else "")
                print(f"    Row {i}: {row_data}")
        
        wb.close()
        
    except Exception as e:
        print(f"Error reading Excel file: {e}")

if __name__ == "__main__":
    check_excel_file() 