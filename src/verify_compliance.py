#!/usr/bin/env python3
"""
Enhanced Script to verify SBOM compliance with baseline requirements
Updated to work with the new comprehensive SBOM structure
"""

import openpyxl
import json
from datetime import datetime
from sbom_exporter import SBOMExporter
from sbom_builder import SBOMBuilder
from highlight_api import HighlightAPI
from config_loader import load_config
from compliance_analyzer import SBOMComplianceAnalyzer
import os
import logging

# Import logging configuration
from logging_config import setup_module_logging

# Set up separated logging for verify_compliance module
logger, log_files = setup_module_logging('verify_compliance')

CONFIG_PATH = 'config/config.json'

def verify_sbom_compliance():
    """Verify the generated SBOM against baseline requirements"""
    
    # Enhanced baseline component information requirements
    baseline_fields = [
        "Component Name",
        "Component Version", 
        "Component Description",
        "Component Supplier",
        "Component License",
        "Component Origin",
        "Component Dependencies",
        "Vulnerabilities",
        "Patch Status",
        "Release Date",
        "End-of-Life (EOL) Date",
        "Criticality",
        "Usage Restrictions",
        "Checksums or Hashes",
        "Comments or Notes",
        "Author of SBOM Data",
        "Timestamp",
        "Executable Property",
        "Archive Property",
        "Structured Property",
        "Unique Identifier"
    ]
    
    print("🔍 Enhanced SBOM Compliance Verification Report")
    print("=" * 60)
    print(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    
    try:
        # Try to find the most recent Excel file in Reports directory
        reports_dir = "Reports"
        if not os.path.exists(reports_dir):
            print("❌ Reports directory not found. Please run the SBOM generator first.")
            return
        
        # Find the most recent Excel file
        excel_files = [f for f in os.listdir(reports_dir) if f.endswith('.xlsx')]
        if not excel_files:
            print("❌ No Excel files found in Reports directory. Please run the SBOM generator first.")
            return
        
        # Sort by modification time and get the most recent
        excel_files.sort(key=lambda x: os.path.getmtime(os.path.join(reports_dir, x)), reverse=True)
        latest_excel_file = os.path.join(reports_dir, excel_files[0])
        
        print(f"📊 Analyzing Excel File: {excel_files[0]}")
        print()
        
        # Load the Excel file
        wb = openpyxl.load_workbook(latest_excel_file)
        
        print(f"📊 Excel File Analysis:")
        print(f"   - Total Worksheets: {len(wb.sheetnames)}")
        print(f"   - Worksheet Names: {wb.sheetnames}")
        print()
        
        # Check Components Complete worksheet
        if 'Components Complete' in wb.sheetnames:
            ws = wb['Components Complete']
            headers = [cell.value for cell in ws[1]]
            
            print("✅ Enhanced Baseline Field Compliance Check:")
            print("-" * 50)
            
            compliance_count = 0
            missing_fields = []
            enhanced_fields = []
            
            for field in baseline_fields:
                if field in headers:
                    print(f"   ✅ {field}")
                    compliance_count += 1
                    enhanced_fields.append(field)
                else:
                    print(f"   ❌ {field} - MISSING")
                    missing_fields.append(field)
            
            print()
            print(f"📈 Enhanced Compliance Summary:")
            print(f"   - Total Required Fields: {len(baseline_fields)}")
            print(f"   - Implemented Fields: {compliance_count}")
            print(f"   - Missing Fields: {len(missing_fields)}")
            print(f"   - Compliance Rate: {(compliance_count/len(baseline_fields)*100):.1f}%")
            
            if missing_fields:
                print()
                print("⚠️  Missing Fields:")
                for field in missing_fields:
                    print(f"   - {field}")
            
            print()
            
            # Enhanced field coverage analysis
            print("🔍 Enhanced Field Coverage Analysis:")
            print("-" * 40)
            
            # Check for additional fields beyond baseline
            additional_fields = [h for h in headers if h not in baseline_fields]
            if additional_fields:
                print(f"   ✅ Additional Fields Found: {len(additional_fields)}")
                for field in additional_fields:
                    print(f"      + {field}")
            else:
                print("   ℹ️  No additional fields beyond baseline requirements")
            
            print()
            
            # Check vulnerability data
            if 'Vulnerabilities' in wb.sheetnames:
                ws_vuln = wb['Vulnerabilities']
                vuln_count = ws_vuln.max_row - 1  # Subtract header
                print(f"🔒 Enhanced Vulnerability Analysis:")
                print(f"   - Total Vulnerabilities Found: {vuln_count}")
                
                # Count by severity
                severity_counts = {"CRITICAL": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0}
                for row in range(2, ws_vuln.max_row + 1):
                    severity = ws_vuln.cell(row=row, column=4).value  # Severity column
                    if severity in severity_counts:
                        severity_counts[severity] += 1
                
                for severity, count in severity_counts.items():
                    if count > 0:
                        print(f"   - {severity}: {count}")
            
            print()
            
            # Check security analysis
            if 'Security Analysis' in wb.sheetnames:
                ws_sec = wb['Security Analysis']
                print(f"🛡️  Enhanced Security Analysis:")
                print(f"   - Components with vulnerabilities: {sum(1 for row in range(2, ws_sec.max_row + 1) if int(ws_sec.cell(row=row, column=3).value or 0) > 0)}")
                print(f"   - Components requiring immediate action: {sum(1 for row in range(2, ws_sec.max_row + 1) if int(ws_sec.cell(row=row, column=4).value or 0) > 0)}")
            
            print()
            print("📋 Enhanced SBOM Quality Assessment:")
            print("-" * 40)
            
            # Enhanced quality checks
            quality_score = 0
            total_checks = 7  # Increased from 5 to 7
            
            # Check 1: Has vulnerability data
            if vuln_count > 0:
                print("   ✅ Vulnerability data present")
                quality_score += 1
            else:
                print("   ⚠️  No vulnerability data found")
            
            # Check 2: Has multiple worksheets
            if len(wb.sheetnames) >= 5:
                print("   ✅ Multiple logical worksheets for organization")
                quality_score += 1
            else:
                print("   ⚠️  Limited worksheet organization")
            
            # Check 3: Has security analysis
            if 'Security Analysis' in wb.sheetnames:
                print("   ✅ Security analysis and risk scoring included")
                quality_score += 1
            else:
                print("   ❌ No security analysis worksheet")
            
            # Check 4: Has metadata
            if 'Metadata' in wb.sheetnames:
                print("   ✅ Complete metadata tracking")
                quality_score += 1
            else:
                print("   ❌ Missing metadata worksheet")
            
            # Check 5: Has comprehensive component data
            if ws.max_row > 20:  # More than 20 components
                print("   ✅ Comprehensive component coverage")
                quality_score += 1
            else:
                print("   ⚠️  Limited component coverage")
            
            # Check 6: Enhanced field coverage (new)
            if compliance_count >= 15:  # At least 15 out of 21 fields
                print("   ✅ Excellent field coverage (enhanced)")
                quality_score += 1
            elif compliance_count >= 10:
                print("   ✅ Good field coverage")
                quality_score += 1
            else:
                print("   ⚠️  Limited field coverage")
            
            # Check 7: Additional fields beyond baseline (new)
            if len(additional_fields) >= 3:
                print("   ✅ Rich additional metadata fields")
                quality_score += 1
            elif len(additional_fields) >= 1:
                print("   ✅ Some additional metadata fields")
                quality_score += 1
            else:
                print("   ℹ️  No additional metadata fields")
            
            print()
            print(f"🎯 Enhanced Quality Score: {quality_score}/{total_checks} ({(quality_score/total_checks*100):.1f}%)")
            
            if quality_score >= 6:
                print("🏆 EXCELLENT - SBOM meets high-quality standards with enhanced coverage")
            elif quality_score >= 4:
                print("✅ GOOD - SBOM meets most requirements with good coverage")
            else:
                print("⚠️  NEEDS IMPROVEMENT - SBOM requires enhancements")
            
            print()
            print("🔍 API Integration Assessment:")
            print("-" * 35)
            
            # Assess API integration effectiveness
            api_fields = [
                "Component Origin",
                "Component Dependencies", 
                "Patch Status",
                "Release Date",
                "End-of-Life (EOL) Date",
                "Criticality",
                "Usage Restrictions",
                "Checksums or Hashes",
                "Comments or Notes",
                "Executable Property",
                "Archive Property",
                "Structured Property"
            ]
            
            api_coverage = sum(1 for field in api_fields if field in enhanced_fields)
            api_percentage = (api_coverage / len(api_fields)) * 100
            
            print(f"   - API-Enhanced Fields: {api_coverage}/{len(api_fields)}")
            print(f"   - API Integration Effectiveness: {api_percentage:.1f}%")
            
            if api_percentage >= 80:
                print("   ✅ Excellent API integration - Most fields successfully extracted")
            elif api_percentage >= 60:
                print("   ✅ Good API integration - Many fields successfully extracted")
            elif api_percentage >= 40:
                print("   ⚠️  Moderate API integration - Some fields extracted")
            else:
                print("   ❌ Limited API integration - Few fields extracted")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error analyzing SBOM: {e}")
        logger.error(f"Error in verify_sbom_compliance: {e}")

def verify_json_sbom_compliance():
    """Verify JSON SBOM file compliance"""
    print("🔍 JSON SBOM Compliance Verification")
    print("=" * 50)
    
    try:
        # Find the most recent JSON file
        reports_dir = "Reports"
        if not os.path.exists(reports_dir):
            print("❌ Reports directory not found.")
            return
        
        json_files = [f for f in os.listdir(reports_dir) if f.endswith('.json') and not f.endswith('_cyclonedx.json')]
        if not json_files:
            print("❌ No JSON SBOM files found.")
            return
        
        # Sort by modification time and get the most recent
        json_files.sort(key=lambda x: os.path.getmtime(os.path.join(reports_dir, x)), reverse=True)
        latest_json_file = os.path.join(reports_dir, json_files[0])
        
        print(f"📊 Analyzing JSON File: {json_files[0]}")
        print()
        
        with open(latest_json_file, 'r') as f:
            sbom_data = json.load(f)
        
        components = sbom_data.get('components', [])
        if not components:
            print("❌ No components found in JSON SBOM")
            return
        
        print(f"📊 JSON SBOM Analysis:")
        print(f"   - Total Components: {len(components)}")
        print(f"   - SBOM Version: {sbom_data.get('sbomVersion', 'Unknown')}")
        print()
        
        # Analyze field coverage in JSON
        field_coverage = {}
        for component in components:
            for key, value in component.items():
                if key not in field_coverage:
                    field_coverage[key] = 0
                if value and value != "Unknown" and value != "Unavailable from CAST Highlight":
                    field_coverage[key] += 1
        
        print("📈 JSON Field Coverage:")
        for field, count in sorted(field_coverage.items()):
            percentage = (count / len(components)) * 100
            print(f"   - {field}: {count}/{len(components)} ({percentage:.1f}%)")
        
        # Check properties coverage
        properties_coverage = {}
        for component in components:
            for prop in component.get('properties', []):
                prop_name = prop.get('name', 'unknown')
                if prop_name not in properties_coverage:
                    properties_coverage[prop_name] = 0
                properties_coverage[prop_name] += 1
        
        if properties_coverage:
            print()
            print("🔧 Properties Coverage:")
            for prop_name, count in sorted(properties_coverage.items()):
                percentage = (count / len(components)) * 100
                print(f"   - {prop_name}: {count}/{len(components)} ({percentage:.1f}%)")
        
    except Exception as e:
        print(f"❌ Error analyzing JSON SBOM: {e}")
        logger.error(f"Error in verify_json_sbom_compliance: {e}")

if __name__ == "__main__":
    print("🚀 Starting Enhanced SBOM Compliance Verification")
    print()
    
    # Verify Excel SBOM
    verify_sbom_compliance()
    
    print()
    print("=" * 60)
    print()
    
    # Verify JSON SBOM
    verify_json_sbom_compliance()
    
    print()
    print("✅ Enhanced SBOM Compliance Verification Complete")